from openpyxl import load_workbook
import openpyxl

def read_all_data(path):
    workbook = load_workbook(filename=path)
    # print(f'Worksheet names: {workbook.sheetnames}')
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f'Title = {sheet.title}')
        for row in sheet.rows:
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                print(f'{cell.column_letter}{cell.row} = {cell.value}')

if __name__ == '__main__':
    read_all_data('books.xlsx')