from openpyxl import load_workbook

def iterating_col(path, sheet_name, col):
    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f'{sheet_name} not found.')
        return
    sheet = workbook[sheet_name]

    for cell in sheet[col]:
        print(f'{cell.column_letter}{cell.row} = {cell.value}')


if __name__ == '__main__':
    iterating_col('books.xlsx', 'sales', 'B')