from openpyxl import load_workbook

def get_sheet_info(path, sheet_name):
    workbook = load_workbook(filename=path)
    # print(f'Worksheet names: {workbook.sheetnames}')

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f'The title of this worksheet is: {sheet.title}')
        print(f'Sheet that contain data: {sheet.calculate_dimension()}')

if __name__ == '__main__':
    get_sheet_info('books.xlsx', sheet_name='sales')