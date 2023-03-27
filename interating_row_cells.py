from openpyxl import load_workbook

def iterating_over_values(path, sheet_name, col):
    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f'{sheet_name} not found.')
        return
    sheet = workbook[sheet_name]

    for value in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=4, values_only=True):
        print(value)

if __name__ == '__main__':
    iterating_over_values('books.xlsx', 'sales', 'B')